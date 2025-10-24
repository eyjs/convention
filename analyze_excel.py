import openpyxl
import os

def analyze_excel(file_path):
    print(f"\n{'='*80}")
    print(f"Analyzing: {os.path.basename(file_path)}")
    print(f"{'='*80}\n")

    wb = openpyxl.load_workbook(file_path, data_only=False)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"[Sheet: {sheet_name}]")
        print(f"  Max Row: {ws.max_row}, Max Column: {ws.max_column}\n")

        # 헤더 출력 (첫 3행)
        print("  Headers:")
        for row_idx in range(1, min(4, ws.max_row + 1)):
            print(f"    Row {row_idx}:")
            for col_idx in range(1, min(11, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                value = str(cell.value).strip() if cell.value else ""

                # 줄바꿈 감지
                has_linebreak = "\n" in value if value else False
                linebreak_indicator = " [LF]" if has_linebreak else ""

                # 값 표시 (길면 자르기)
                display_value = value[:30] + "..." if len(value) > 30 else value
                display_value = display_value.replace("\n", "\\n")

                if display_value:
                    print(f"      [{col_idx}] = {display_value}{linebreak_indicator}")
            print()

        # 샘플 데이터 (4-7행)
        if ws.max_row >= 4:
            print("  Sample Data:")
            for row_idx in range(4, min(8, ws.max_row + 1)):
                print(f"    Row {row_idx}:")
                has_data = False
                for col_idx in range(1, min(11, ws.max_column + 1)):
                    cell = ws.cell(row_idx, col_idx)
                    value = str(cell.value).strip() if cell.value else ""

                    if value:
                        has_data = True
                        display_value = value[:30] + "..." if len(value) > 30 else value
                        display_value = display_value.replace("\n", "\\n")
                        print(f"      [{col_idx}] = {display_value}")

                if not has_data:
                    print(f"      (empty)")
                print()

        print()

# 샘플 파일들 분석
sample_dir = "Sample"
files = [
    "참석자업로드_샘플.xlsx",
    "일정업로드_샘플.xlsx",
    "속성업로드_샘플.xlsx"
]

for file in files:
    file_path = os.path.join(sample_dir, file)
    if os.path.exists(file_path):
        analyze_excel(file_path)
    else:
        print(f"File not found: {file_path}")
